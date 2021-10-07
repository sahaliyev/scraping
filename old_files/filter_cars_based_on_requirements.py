from openpyxl import load_workbook
from sqlalchemy import create_engine
import pandas as pd
from datetime import datetime
import os
from turbo_db_operations import TurboDbOperations


class FilterCars(TurboDbOperations):
    @staticmethod
    def engine_creator():
        database = {
            'local': {
                'NAME': 'turbo',
                'USER': 'postgres',
                'PASSWORD': 'postgres',
                'HOST': 'localhost',
                'PORT': 5432,
            },
            'production': {
                'NAME': 'turbo',
                'USER': 'postgres',
                'PASSWORD': 'temp1Parol8#',
                'HOST': 'localhost',
                'PORT': 5432,
            }
        }
        db = database['local']
        engine_string = "postgresql+psycopg2://{user}:{password}@{host}:{port}/{database}".format(
            user=db['USER'],
            password=db['PASSWORD'],
            host=db['HOST'],
            port=db['PORT'],
            database=db['NAME'],
        )

        return create_engine(engine_string)

    @staticmethod
    def str_price_to_int_converter(price_str):
        price_str = price_str.replace(" ", '')
        return int(price_str)

    @staticmethod
    def extract_engine_size(engine):
        return '1.4' in engine

    @staticmethod
    def filter_df_marka(df, marka):
        return df[(df['marka'] == marka)]

    @staticmethod
    def filter_df_model(df, model):
        return df[(df['model'] == model)]

    @staticmethod
    def filter_df_year(df, year):
        min_year, max_year = year
        return df[(df['year'].between(min_year, max_year))]

    def filter_df_price(self, df, price):
        min_price, max_price = price
        return df[(df['price'].apply(self.str_price_to_int_converter).between(min_price, max_price))]

    @staticmethod
    def filer_df_gearbox(df, gearbox):
        if type(gearbox) == str:
            return df[(df['gearbox'] == gearbox)]
        elif type(gearbox) == list:
            return df[(df['gearbox'].isin(gearbox))]

    @staticmethod
    def filter_df_bay_type(df, ban_type):
        if type(ban_type) == str:
            return df[(df['ban_type'] == ban_type)]
        elif type(ban_type) == list:
            return df[(df['gearbox'].isin(ban_type))]

    @staticmethod
    def filter_df_fuel_type(df, fuel_type):
        if type(fuel_type) == str:
            return df[(df['fuel_type'] == fuel_type)]
        elif type(fuel_type) == list:
            return df[(df['gearbox'].isin(fuel_type))]

    @staticmethod
    def filter_df_keywords(df):
        keywords = ['Təcili', 'Yeni gəlib', 'Öz avtomobilimdi', 'Bakıda sürülməyib',
                    'Pul lazımdır', 'İdeal vəziyyətdədir', 'Servis maşınıdır']

        return df[df['description'].str.contains('|'.join(keywords), case=False)]

    @staticmethod
    def check_file_existence(path):
        return os.path.isfile(path)

    def filter_df_engine(self, df):
        return df[(df['engine'].apply(self.extract_engine_size))]

    def filter_dataframe_without_keyword(self, df, **item):
        today = datetime.now().strftime("%d.%m.%Y")

        if item.get('marka'):
            df = self.filter_df_marka(df, item.get('marka'))
        if item.get('model'):
            df = self.filter_df_model(df, item.get('model'))
        if item.get('year'):
            df = self.filter_df_year(df, item.get('year'))
        if item.get('price'):
            df = self.filter_df_price(df, item.get('price'))
        if item.get('gearbox'):
            df = self.filer_df_gearbox(df, item.get('gearbox'))
        if item.get('ban_type'):
            df = self.filter_df_bay_type(df, item.get('ban_type'))
        if item.get('fuel_type'):
            df = self.filter_df_fuel_type(df, item.get('fuel_type'))

        specific_rows = [*item]
        specific_rows.insert(5, 'currency')  # to add currency
        specific_rows = specific_rows[1:]
        df_with_specific_rows = df[specific_rows]

        file_name = f"{today}_results_without_keyword.xlsx"
        file_path = f"excels/{file_name}"

        marka = model = 'Missing'
        if item.get('marka'):
            marka = item.get('marka')
        if item.get('model'):
            model = item.get('model')

        if self.check_file_existence(file_path):
            book = load_workbook(file_path)
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                writer.book = book
                df_with_specific_rows.to_excel(writer, sheet_name=f"{marka} {model}", index=False)
        else:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_with_specific_rows.to_excel(writer, sheet_name=f"{marka} {model}", index=False)

    def filter_dataframe_with_keyword(self, df, **item):
        today = datetime.now().strftime("%d.%m.%Y")

        if item.get('marka'):
            df = self.filter_df_marka(df, item.get('marka'))
        if item.get('model'):
            df = self.filter_df_model(df, item.get('model'))
        if item.get('year'):
            df = self.filter_df_year(df, item.get('year'))
        if item.get('price'):
            df = self.filter_df_price(df, item.get('price'))
        if item.get('gearbox'):
            df = self.filer_df_gearbox(df, item.get('gearbox'))
        if item.get('ban_type'):
            df = self.filter_df_bay_type(df, item.get('ban_type'))
        if item.get('fuel_type'):
            df = self.filter_df_fuel_type(df, item.get('fuel_type'))
        df = self.filter_df_keywords(df)

        specific_rows = [*item]
        specific_rows.insert(5, 'currency')
        specific_rows.append('description')
        df_with_specific_rows = df[specific_rows]

        file_name = f"{today}_results_with_keywords.xlsx"
        file_path = f"excels/{file_name}"

        marka = model = 'Missing'
        if item.get('marka'):
            marka = item.get('marka')
        if item.get('model'):
            model = item.get('model')

        if self.check_file_existence(file_path):
            book = load_workbook(file_path)
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                writer.book = book
                df_with_specific_rows.to_excel(writer, sheet_name=f"{marka} {model}", index=False)
        else:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_with_specific_rows.to_excel(writer, sheet_name=f"{marka} {model}", index=False)

    def main(self):
        engine = self.engine_creator()
        df = pd.read_sql_table('car', engine)
        requirements = self.get_requirements()
        for item in requirements:
            self.filter_dataframe_with_keyword(df, **item)
            self.filter_dataframe_without_keyword(df, **item)


if __name__ == '__main__':
    FilterCars().main()
